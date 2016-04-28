#!/usr/bin/python2.7
from __future__ import division
from openpyxl import Workbook
from openpyxl import load_workbook
from nltk.corpus import stopwords
from nltk.tokenize import sent_tokenize
from nltk import pos_tag
from nltk import word_tokenize
from nltk.stem import WordNetLemmatizer
from nltk.corpus import wordnet as wn
from collections import Counter
from math import log
from decimal import Decimal
from fileinput import input
from operator import itemgetter
from openpyxl import Workbook
from shutil import rmtree
from time import time
from math import floor
import os
import re
import sys

INPUT='texts'
sys.stdout.write("Loading tf-idf values...\r")
sys.stdout.flush()
tfidf_wb = load_workbook(filename='tfidf-final.xlsx')
sys.stdout.write("Loading tf-idf values complete\n")
sys.stdout.flush()
wb = Workbook()
stopword_list = []

def dir_function(sub_folder):
    files = []
    for fil in os.listdir("./"+sub_folder):
        if fil.endswith(".txt"):
            files.append(fil)
    return files

def clear_stopword_list():
    global stopword_list
    del stopword_list[:]

def load_stopwords_list(sheetname):
    clear_stopword_list()
    ws = tfidf_wb[sheetname]
    start_index = int(floor(0.75*(ws.max_row-1))+1)
    for i in range(1,start_index):
        stopword_list.append(ws.cell(row=i, column = 1).value)

def getWordNetType(tag):
    if tag in ['JJ', 'JJR', 'JJS']:
        return wn.ADJ
    elif tag in ['NN', 'NNS', 'NNP', 'NNPS','POS','FW']:
        return wn.NOUN
    elif tag in ['RB', 'RBR', 'RBS','WRB']:
        return wn.ADV
    elif tag in ['VB', 'VBD', 'VBG', 'VBN', 'VBP', 'VBZ']:
        return wn.VERB
    return wn.NOUN

def pos_tagging(input):
    global cnt
    wn_lem = WordNetLemmatizer()
    details = pos_tag(word_tokenize(input))
    new_details = []
    for word, typ in details:
        word = wn_lem.lemmatize(unicode(word.lower(),errors='replace'),pos=getWordNetType(typ))
        if word == typ:
            continue
        elif type in [':','$','(',')','--','.','*']:
            continue
        elif word in ["'t","'ve","'d","'ll","'s","'m","'am","n't","'re"]:
            continue
        elif word in stopword_list:
            continue
        else:
            new_details.append(word)
    return new_details

def process_reference(input,sheetname):
    global wb
    words = []
    ws = wb.create_sheet(sheetname)
    count = 2
    ws.cell(row=1,column=1).value='Sentence'
    ws.cell(row=1,column=2).value='Keywords'
    ws.cell(row=1,column=3).value='Label'
    sentences = sent_tokenize(input)
    for sentence in sentences:
        sentence1 = re.sub('[^A-Za-z0-9_ ]+',' ',sentence)
        sentence1 = re.sub(' +',' ',sentence1)
        lis = pos_tagging(sentence1)
        ws.cell(row=count, column=1).value=sentence
        ws.cell(row=count, column=2).value=' '.join(lis)
        count +=1

def read_Documents(source,sheetname):
    data = ''
    with open(source,'r') as f:
        for temp in f.readlines():
            temp = temp.replace('\r\n',' ')
            data = data + temp
        data = re.sub("[^A-Za-z0-9_. ]+",' ',data)
        data = re.sub(' +',' ',data)
    process_reference(data,sheetname)

def init():
    start_ts = time()
    dir_list = dir_function(INPUT)
    file_count = 0
    total_count= len(dir_list)
    sys.stdout.write("Processing text documents: %.2f%% complete\r"%float(file_count*100.0/total_count))
    sys.stdout.flush()
    for filename in dir_list:
        file_count +=1
        sheetname = filename.strip('.txt')[:30]
        load_stopwords_list(sheetname)
        read_Documents(INPUT+'/'+filename,sheetname)
        sys.stdout.write("Processing text documents: %.2f%% complete\r"%float(file_count*100.0/total_count))
        sys.stdout.flush()
    sys.stdout.write("\nSaving final result as result.xlsx ...\r")
    sys.stdout.flush()
    wb.save('result.xlsx')
    sys.stdout.write("Saving final result as result.xlsx complete\n")
    sys.stdout.flush()
    time_taken = divmod((time()-start_ts),60)
    print("Overall time taken: %d minutes and %d seconds" %(time_taken[0],time_taken[1]))

if __name__ == '__main__':
    init()
