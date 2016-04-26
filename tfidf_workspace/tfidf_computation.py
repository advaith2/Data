#!/usr/bin/python2.7
from __future__ import division
from openpyxl import Workbook
from nltk.corpus import stopwords
from nltk.tokenize import sent_tokenize
from nltk import pos_tag
from nltk import word_tokenize
from nltk.stem import WordNetLemmatizer
from nltk.corpus import wordnet as wn
from collections import Counter
from math import log
from decimal import Decimal
from fileinput import input
from operator import itemgetter
from openpyxl import Workbook
from shutil import rmtree
from time import time
import os
import re
import sys

cnt = Counter()
idf = dict()
wb = Workbook()
WORD_COUNT = 'wordcounts'
INPUT='texts'
TF_IDF='tfidf'
WB_NAME='tfidf-final.xlsx'

def getWordNetType(tag):
    if tag in ['JJ', 'JJR', 'JJS']:
        return wn.ADJ
    elif tag in ['NN', 'NNS', 'NNP', 'NNPS','POS','FW']:
        return wn.NOUN
    elif tag in ['RB', 'RBR', 'RBS','WRB']:
        return wn.ADV
    elif tag in ['VB', 'VBD', 'VBG', 'VBN', 'VBP', 'VBZ']:
        return wn.VERB
    return wn.NOUN

def pos_tagging(input):
    global cnt
    wn_lem = WordNetLemmatizer()
    details = pos_tag(word_tokenize(input))
    new_details = []
    for word, typ in details:
        if word == typ:
            continue
        elif type in [':','$','(',')','--','.','*']:
            continue
        elif word in ["'t","'ve","'d","'ll","'s","'m","'am","n't","'re"]:
            continue
        else:
            word1 = wn_lem.lemmatize(unicode(word.lower(),errors='replace'),pos=getWordNetType(typ))
            cnt[word1] += 1

def process_word_count(filename):
    data = ''
    with open(filename,'r') as f:
        for temp in f.readlines():
            temp = temp.replace('\r\n',' ')
            data = data + temp
        data = re.sub("[^\w']",' ',data)
        data = re.sub(' +',' ',data)
    sentences = sent_tokenize(data)
    for sentence in sentences:
        sentence = re.sub('[^A-Za-z0-9 ]+',' ',sentence)
        sentence = re.sub(' +',' ',sentence)
        pos_tagging(sentence)

def term_frequency():
    global cnt
    file_count = 0
    dir_list = dir_function(INPUT)
    total_count = len(dir_list)
    check_and_create_dir(WORD_COUNT)
    sys.stdout.write("Computing Term Frequency: %.2f%% complete\r"%float(file_count*100.0/total_count))
    sys.stdout.flush()
    for filename in dir_list:
        file_count +=1
        process_word_count(INPUT+'/'+filename)
        l1 = list(cnt)
        w_count = len(l1)
        with open(WORD_COUNT+'/'+filename,'w') as f1:
            for word, count in cnt.items():
                word1 = word+' '+str(count/w_count)
                f1.write(word1.encode('utf-8').strip()+'\n')
        for word in l1:
            del cnt[word]
        sys.stdout.write("Computing Term Frequency: %.2f%% complete\r"%float(file_count*100.0/total_count))
        sys.stdout.flush()

def update_idf(filename):
    global cnt
    with open(filename,'r') as f:
        for temp in f.readlines():
            words = temp.split()
            if words:
                cnt[words[0]] += 1

def inverse_doc_frequency():
    sys.stdout.write("\nComputing Inverse Document Frequency\r")
    sys.stdout.flush()
    dir_list = dir_function('./'+WORD_COUNT)
    total_count = len(dir_list)
    for filename in dir_list:
        update_idf(WORD_COUNT+'/'+filename)
    with open('idf.txt','w') as f1:
        for word, count in cnt.items():
            word1 = word+' '+str(log(total_count/count))
            f1.write(str(word1)+'\n')
    print 'Computing Inverse Document Frequency Completed'

def load_idf(filename):
    global idf
    with open(filename,'r') as f:
        for line in f.readlines():
            tokens = line.split()
            idf[tokens[0]]=Decimal(tokens[1])

def calculate_tfidf(filename):
    for line in input(filename,inplace='true'):
        tokens = line.split()
        tf = Decimal(tokens[1])
        result = tf*idf[tokens[0]]
        res_str = "%.20f"%result
        print tokens[0]+' '+tokens[1]+' '+res_str+'\n',

def tfidf_init():
    load_idf('idf.txt')
    file_count = 0
    dir_list = dir_function('./'+WORD_COUNT)
    total_count = len(dir_list)
    sys.stdout.write("Computing tfidf: %.2f%% complete\r"%float(file_count*100.0/total_count))
    sys.stdout.flush()
    for filename in dir_list:
        file_count +=1
        calculate_tfidf(WORD_COUNT+'/'+filename)
        sys.stdout.write("Computing tfidf: %.2f%% complete\r"%float(file_count*100.0/total_count))
        sys.stdout.flush()
    print

def sort_file (fileName,dest):
    lis = []
    with open(fileName,'r') as f:
        for lines in f.readlines():
            tokens = lines.split()
            temp = (tokens[0],Decimal(tokens[1]),Decimal(tokens[2]))
            lis.append(temp)
    lis.sort(key=itemgetter(2))
    with open(dest,'w') as f1:
        for l in lis:
            f1.write(l[0]+' '+'%.20f'%l[1]+' '+'%.20f'%l[2]+'\n')

def sorting_init():
    check_and_create_dir(TF_IDF)
    file_count = 0
    dir_list = dir_function('./'+WORD_COUNT)
    total_count= len(dir_list)
    sys.stdout.write("Sorting files: %.2f%% complete\r"%float(file_count*100.0/total_count))
    sys.stdout.flush()
    for filename in dir_list:
        file_count +=1
        sort_file(WORD_COUNT+'/'+filename,TF_IDF+'/'+filename)
        sys.stdout.write("Sorting files: %.2f%% complete\r"%float(file_count*100.0/total_count))
        sys.stdout.flush()
    print

def addSheet(src_folder,fileName):
    global wb
    ws = wb.create_sheet()
    ws.title=fileName.strip('.txt')[:30]
    count = 2
    ws.cell(row=1,column=1).value='Word'
    ws.cell(row=1,column=2).value='Term-Frequency'
    ws.cell(row=1,column=3).value='tf-idf'
    with open(src_folder+'/'+fileName,'r') as f:
        for lines in f.readlines():
            tokens = lines.split()
            ws.cell(row = count, column=1).value=tokens[0]
            ws.cell(row = count, column=2).value=tokens[1]
            ws.cell(row = count, column=3).value=tokens[2]
            count +=1

def sheet_creation_init():
    global wb
    file_count = 0
    dir_list = dir_function('./'+TF_IDF)
    total_count= len(dir_list)
    sys.stdout.write("Creating excel sheet: %.2f%% complete\r"%float(file_count*100.0/total_count))
    sys.stdout.flush()
    for filename in dir_list:
        file_count +=1
        addSheet(TF_IDF,filename)
        sys.stdout.write("Creating excel sheet: %.2f%% complete\r"%float(file_count*100.0/total_count))
        sys.stdout.flush()
    print
    sys.stdout.write("Saving excel sheet... \r")
    sys.stdout.flush()
    wb.save(WB_NAME)
    sys.stdout.write("Saving excel sheet complete \r")
    sys.stdout.flush()
    print

def delete_folder(folder_name):
    rmtree(folder_name)

def clear_counter():
    global cnt
    for word in list(cnt):
        del cnt[word]

def dir_function(sub_folder):
    files = []
    for fil in os.listdir("./"+sub_folder):
        if fil.endswith(".txt"):
            files.append(fil)
    return files

def check_and_create_dir(dir_name):
    if not os.path.exists(dir_name):
        os.makedirs(dir_name)

def global_init():
    start_ts = time()
    term_frequency()
    inverse_doc_frequency()
    tfidf_init()
    sorting_init()
    sheet_creation_init()
    sys.stdout.write("Deleting temp folders...\r")
    sys.stdout.flush()
    delete_folder(WORD_COUNT)
    delete_folder(TF_IDF)
    sys.stdout.write("Deleting temp folders complete\r")
    sys.stdout.flush()
    time_taken = divmod((time()-start_ts),60)
    print("\nOverall time taken: %d minutes and %d seconds" %(time_taken[0],time_taken[1]))

if __name__ == '__main__':
    global_init()
